"""
ACCESS (access.trade.gov) CVD 케이스 파일 자동 다운로드 + 엑셀 저장
- 원래 안정적으로 작동하던 방식
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import os, time, logging, openpyxl, json, threading
from openpyxl.styles import Font, PatternFill, Alignment

# ────────────────────────────────────────────
# 설정
# ────────────────────────────────────────────

DOWNLOAD_DIR    = os.path.join(os.path.expanduser("~"), "Downloads", "ACCESS_CVD")
EXCEL_PATH      = os.path.join(DOWNLOAD_DIR, "ACCESS_CVD_목록.xlsx")
CHECKPOINT_PATH = os.path.join(DOWNLOAD_DIR, "checkpoint.json")

CASE_NUMBERS = [
    "C-580-602", "C-580-818", "C-580-835", "C-580-837",
    "C-580-842", "C-580-849", "C-580-851", "C-580-857",
    "C-580-862", "C-580-866", "C-580-869", "C-580-873",
    "C-580-875", "C-580-877", "C-580-879", "C-580-882",
    "C-580-884", "C-580-888", "C-580-898", "C-580-910",
    "C-580-913", "C-580-917", "C-580-920", "C-580-989",
]

BASE_URL   = "https://access.trade.gov"
SEARCH_URL = f"{BASE_URL}/search"

CLICK_WAIT  = 0.8
POPUP_WAIT  = 2.5
PAGE_WAIT   = 5.0
RETRY_COUNT = 3

HEADERS = ["Filed Date", "Case Number", "Segment", "Type", "Title",
           "Filed By", "Pertaining To", "Security", "Barcode"]

# ────────────────────────────────────────────
# 로깅
# ────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("access_download.log", encoding="utf-8"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

# ────────────────────────────────────────────
# 엑셀
# ────────────────────────────────────────────

def init_excel():
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CVD 목록"
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center")
        widths = [12, 13, 25, 20, 40, 25, 25, 12, 15]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        wb.save(EXCEL_PATH)
    return wb, ws

def append_to_excel(wb, ws, row_data):
    try:
        ws.append(row_data)
        wb.save(EXCEL_PATH)
    except Exception as e:
        log.warning(f"엑셀 저장 오류: {e}")

# ────────────────────────────────────────────
# 체크포인트
# ────────────────────────────────────────────

def load_checkpoint():
    if os.path.exists(CHECKPOINT_PATH):
        with open(CHECKPOINT_PATH, 'r') as f:
            return json.load(f)
    return {}

def save_checkpoint(case_number, page):
    cp = load_checkpoint()
    cp[case_number] = page
    with open(CHECKPOINT_PATH, 'w') as f:
        json.dump(cp, f)

def is_case_done(case_number):
    return load_checkpoint().get(case_number) == "done"

def get_start_page(case_number):
    val = load_checkpoint().get(case_number, 1)
    return val if val != "done" else 1

def mark_case_done(case_number):
    cp = load_checkpoint()
    cp[case_number] = "done"
    with open(CHECKPOINT_PATH, 'w') as f:
        json.dump(cp, f)

# ────────────────────────────────────────────
# 중복 체크
# ────────────────────────────────────────────

def is_already_downloaded(barcode):
    if not barcode:
        return False
    for filename in os.listdir(DOWNLOAD_DIR):
        name = os.path.splitext(filename)[0]
        if ' (' in name:
            name = name[:name.rfind(' (')]
        if name.strip() == barcode.strip():
            return True
    return False

# ────────────────────────────────────────────
# 드라이버
# ────────────────────────────────────────────

def init_driver():
    options = Options()
    options.add_experimental_option("prefs", {
        "download.default_directory": DOWNLOAD_DIR,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "plugins.always_open_pdf_externally": True,
    })
    driver = webdriver.Chrome(options=options)
    driver.maximize_window()
    return driver

# ────────────────────────────────────────────
# 로그인
# ────────────────────────────────────────────

def wait_for_manual_login(driver):
    driver.get(BASE_URL)
    log.info("로그인 후 Enter를 눌러주세요...")
    input(">>> 로그인 완료 후 Enter: ")
    log.info("자동 다운로드 시작합니다.")

# ────────────────────────────────────────────
# 세션 감시
# ────────────────────────────────────────────

session_watcher_running = False

def session_watcher_thread(driver):
    global session_watcher_running
    while session_watcher_running:
        try:
            stay_btns = driver.find_elements(
                By.XPATH, "//button[contains(text(),'Stay Logged In')] | //a[contains(text(),'Stay Logged In')]"
            )
            for btn in stay_btns:
                if btn.is_displayed():
                    driver.execute_script("arguments[0].click();", btn)
                    log.info("Stay Logged In 버튼 자동 클릭!")
        except:
            pass
        time.sleep(20)

def start_session_watcher(driver):
    global session_watcher_running
    session_watcher_running = True
    t = threading.Thread(target=session_watcher_thread, args=(driver,), daemon=True)
    t.start()
    log.info("세션 감시 스레드 시작")

def stop_session_watcher():
    global session_watcher_running
    session_watcher_running = False

# ────────────────────────────────────────────
# 세션 확인
# ────────────────────────────────────────────

def check_and_restore_session(driver):
    try:
        if "SignIn" in driver.current_url or ("Sign In" in driver.page_source and "Sign Out" not in driver.page_source):
            log.warning("세션 만료 감지! 다시 로그인해주세요.")
            driver.get(BASE_URL)
            input(">>> 로그인 완료 후 Enter: ")
            log.info("재로그인 완료.")
            return True
        return False
    except:
        return False

# ────────────────────────────────────────────
# 팝업 처리
# ────────────────────────────────────────────

def handle_popup(driver, main_window):
    try:
        end_time = time.time() + 1.5
        while time.time() < end_time:
            if len(driver.window_handles) > 1:
                break
            time.sleep(0.15)
        if len(driver.window_handles) > 1:
            for w in driver.window_handles:
                if w != main_window:
                    try:
                        driver.switch_to.window(w)
                        time.sleep(POPUP_WAIT)
                        driver.close()
                    except:
                        pass
            driver.switch_to.window(main_window)
        else:
            time.sleep(POPUP_WAIT)
    except:
        try:
            driver.switch_to.window(main_window)
        except:
            pass

# ────────────────────────────────────────────
# 테이블 데이터 추출
# ────────────────────────────────────────────

def extract_table_rows(driver):
    rows_data = []
    try:
        rows = driver.find_elements(By.XPATH, "//table//tbody//tr")
        for row in rows:
            cols = row.find_elements(By.XPATH, ".//td")
            if len(cols) >= 9:
                rows_data.append([
                    cols[2].text.strip(), cols[3].text.strip(), cols[4].text.strip(),
                    cols[5].text.strip(), cols[6].text.strip(), cols[7].text.strip(),
                    cols[8].text.strip(),
                    cols[9].text.strip() if len(cols) > 9 else "",
                    cols[10].text.strip() if len(cols) > 10 else ""
                ])
    except Exception as e:
        log.warning(f"테이블 추출 오류: {e}")
    return rows_data

# ────────────────────────────────────────────
# 케이스 검색
# ────────────────────────────────────────────

def search_case(driver, case_number):
    check_and_restore_session(driver)
    driver.get(SEARCH_URL)
    wait = WebDriverWait(driver, 20)
    time.sleep(PAGE_WAIT)
    check_and_restore_session(driver)

    search_input = wait.until(EC.presence_of_element_located((By.NAME, "model.SearchKeyword")))
    search_input.clear()
    search_input.send_keys(case_number)

    Select(driver.find_element(By.NAME, "model.SearchType")).select_by_value("Case Number")

    search_btn = driver.find_element(By.XPATH, "//button[contains(text(),'Quick Search')]")
    driver.execute_script("arguments[0].click();", search_btn)
    time.sleep(PAGE_WAIT)

# ────────────────────────────────────────────
# 현재 페이지 처리
# ────────────────────────────────────────────

def process_current_page(driver, case_number, page, wb, ws):
    check_and_restore_session(driver)
    time.sleep(1.5)
    main_window = driver.current_window_handle

    rows_data = extract_table_rows(driver)
    for row_data in rows_data:
        append_to_excel(wb, ws, row_data)
    log.info(f"[{case_number}] 페이지 {page}: 엑셀 {len(rows_data)}행 저장")

    btns = driver.find_elements(By.XPATH, "//i[contains(@class,'bi-cloud-download-fill')]")
    total = len(btns)
    log.info(f"[{case_number}] 페이지 {page}: {total}개 파일")

    count = 0
    for idx in range(total):
        try:
            rows = driver.find_elements(By.XPATH, "//table//tbody//tr")
            barcode = rows[idx].find_elements(By.XPATH, ".//td")[10].text.strip() if len(rows) > idx else ""
            if barcode and is_already_downloaded(barcode):
                log.info(f"[{case_number}] {idx+1}/{total} 스킵 ({barcode})")
                count += 1
                continue
        except:
            pass

        try:
            rows = driver.find_elements(By.XPATH, "//table//tbody//tr")
            if len(rows) > idx:
                pdf_icon = rows[idx].find_elements(By.XPATH, ".//i[contains(@class,'bi-file-pdf')]")
                if not pdf_icon:
                    log.info(f"[{case_number}] {idx+1}/{total} HTML 스킵 ({barcode})")
                    count += 1
                    continue
        except:
            pass

        for attempt in range(RETRY_COUNT):
            try:
                main_window = driver.current_window_handle
                btns = driver.find_elements(By.XPATH, "//i[contains(@class,'bi-cloud-download-fill')]")
                if idx >= len(btns):
                    break
                driver.execute_script("arguments[0].click();", btns[idx])
                time.sleep(CLICK_WAIT)
                handle_popup(driver, main_window)
                log.info(f"[{case_number}] {idx+1}/{total} 완료")
                count += 1
                break
            except Exception as e:
                log.warning(f"[{case_number}] {idx+1} 시도{attempt+1} 실패")
                try:
                    driver.switch_to.window(main_window)
                except:
                    pass
                time.sleep(2)
    return count

# ────────────────────────────────────────────
# 다음 페이지 이동
# ────────────────────────────────────────────

def go_to_next_page(driver, current_page):
    try:
        for btn in driver.find_elements(By.XPATH, "//a[contains(@class,'btn')]"):
            if btn.text.strip() == str(current_page + 1):
                driver.execute_script("arguments[0].scrollIntoView(true);", btn)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", btn)
                time.sleep(PAGE_WAIT)
                return True

        for btn in driver.find_elements(By.XPATH, "//a[.//i[contains(@class,'bi-skip-forward-fill')]]"):
            cls = btn.get_attribute("class") or ""
            if "disabled" not in cls:
                driver.execute_script("arguments[0].scrollIntoView(true);", btn)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", btn)
                time.sleep(PAGE_WAIT)
                return True

        return False
    except:
        return False

# ────────────────────────────────────────────
# 케이스 전체 처리
# ────────────────────────────────────────────

def process_case(driver, case_number, wb, ws):
    if is_case_done(case_number):
        log.info(f"[{case_number}] 이미 완료 - 스킵")
        return 0

    start_page = get_start_page(case_number)

    for attempt in range(RETRY_COUNT):
        try:
            search_case(driver, case_number)
            break
        except Exception as e:
            log.error(f"[{case_number}] 검색 오류 시도{attempt+1}: {e}")
            if attempt == RETRY_COUNT - 1:
                return 0
            time.sleep(3)

    if start_page > 1:
        log.info(f"[{case_number}] 체크포인트: {start_page}페이지부터 재개")
        for p in range(1, start_page):
            if not go_to_next_page(driver, p):
                break
            time.sleep(PAGE_WAIT)

    total = 0
    page = start_page
    while True:
        log.info(f"[{case_number}] 페이지 {page} 처리 중...")
        save_checkpoint(case_number, page)
        total += process_current_page(driver, case_number, page, wb, ws)

        if not go_to_next_page(driver, page):
            log.info(f"[{case_number}] 완료 ({page}페이지)")
            mark_case_done(case_number)
            break

        page += 1

    return total

# ────────────────────────────────────────────
# 메인
# ────────────────────────────────────────────

def main():
    wb, ws = init_excel()
    driver = init_driver()

    try:
        wait_for_manual_login(driver)
        start_session_watcher(driver)

        total = 0
        for case_number in CASE_NUMBERS:
            log.info(f"\n{'='*50}\n케이스: {case_number}\n{'='*50}")
            count = process_case(driver, case_number, wb, ws)
            total += count
            log.info(f"[{case_number}] 완료: {count}개. 누적: {total}개")

        log.info(f"\n전체 완료. 총 {total}개")
        log.info(f"저장: {DOWNLOAD_DIR}")

    finally:
        stop_session_watcher()
        input("\n완료. Enter 누르면 종료...")
        driver.quit()

if __name__ == "__main__":
    main()
